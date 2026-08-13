"""
Laravel Sync Service
Modul untuk menarik data dari Laravel API dan menyimpan ke database MySQL lokal Python
"""

import requests
import logging
from typing import Dict, List, Optional
from datetime import datetime
import mysql.connector

from app.config_db import MYSQL_CONFIG

logger = logging.getLogger('LaravelSync')


class LaravelSyncService:
    def __init__(self, laravel_base_url: str = "https://pkkmb.polinela.ac.id", verify_ssl: bool = True):
        """
        Initialize Laravel Sync Service
        
        Args:
            laravel_base_url: Base URL Laravel server (default: http://127.0.0.1:8000)
            verify_ssl: Verify SSL certificates for HTTPS (default: True)
        """
        self.base_url = laravel_base_url.rstrip('/')
        self.verify_ssl = verify_ssl
        self.session = requests.Session()
        self.session.headers.update({
            'Accept': 'application/json, text/plain, */*',
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })
        
        # Disable SSL verification warning if needed
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
    def _get_conn(self):
        """Dapatkan koneksi MySQL"""
        return mysql.connector.connect(**MYSQL_CONFIG)
    
    def _execute(self, query, params=None, fetch_one=False, fetch_all=False):
        """Execute MySQL query dengan error handling"""
        conn = self._get_conn()
        cursor = conn.cursor()
        
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            result = None
            if fetch_one:
                row = cursor.fetchone()
                if row:
                    columns = [desc[0] for desc in cursor.description]
                    result = dict(zip(columns, row))
            elif fetch_all:
                rows = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                result = [dict(zip(columns, row)) for row in rows]
            else:
                conn.commit()
                result = cursor.lastrowid
            
            return result
        except mysql.connector.Error as e:
            logger.error(f"Database error: {e}")
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()
    
    def test_connection(self) -> Dict:
        """
        Test koneksi ke Laravel API
        
        Returns:
            Dict dengan status koneksi
        """
        try:
            # Try to connect to the sync status endpoint
            response = self.session.get(
                f"{self.base_url}/api/sync/status", 
                timeout=10,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            return {
                'success': True,
                'message': 'Koneksi ke Laravel API berhasil',
                'data': response.json()
            }
        except requests.exceptions.SSLError as e:
            logger.error(f"SSL error to {self.base_url}: {e}")
            return {
                'success': False,
                'message': f'SSL Error: Sertifikat tidak valid. Coba gunakan HTTP atau hubungi administrator server.'
            }
        except requests.exceptions.Timeout:
            logger.error(f"Connection timeout to {self.base_url}")
            return {
                'success': False,
                'message': f'Timeout: Server tidak merespon dalam 10 detik. Pastikan Laravel server berjalan di {self.base_url}'
            }
        except requests.exceptions.ConnectionError as e:
            logger.error(f"Connection error to {self.base_url}: {e}")
            # Check if it's a "Remote end closed connection" error
            error_msg = str(e).lower()
            if 'remote end closed' in error_msg or 'connection aborted' in error_msg:
                return {
                    'success': False,
                    'message': f'Server menutup koneksi: Pastikan endpoint /api/sync/status tersedia di {self.base_url}'
                }
            return {
                'success': False,
                'message': f'Gagal terhubung ke Laravel: Pastikan server berjalan di {self.base_url} dan dapat diakses'
            }
        except requests.exceptions.HTTPError as e:
            logger.error(f"HTTP error from {self.base_url}: {e}")
            status_code = e.response.status_code
            if status_code == 404:
                return {
                    'success': False,
                    'message': f'Endpoint tidak ditemukan (404): Pastikan route /api/sync/status sudah terdaftar di Laravel'
                }
            elif status_code == 500:
                return {
                    'success': False,
                    'message': f'Server error (500): Ada error di Laravel. Cek log Laravel untuk detail.'
                }
            return {
                'success': False,
                'message': f'Error HTTP {status_code}: {e.response.text[:200]}'
            }
        except requests.exceptions.RequestException as e:
            logger.error(f"Request error to {self.base_url}: {e}")
            return {
                'success': False,
                'message': f'Gagal terhubung ke Laravel: {str(e)}'
            }
    
    def fetch_mahasiswa(self) -> Dict:
        """
        Menarik semua data mahasiswa dari Laravel API
        
        Returns:
            Dict dengan status dan data mahasiswa
        """
        try:
            logger.info("Fetching mahasiswa data from Laravel...")
            response = self.session.get(
                f"{self.base_url}/api/sync/mahasiswa", 
                timeout=60,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            
            data = response.json()
            if not data.get('success'):
                return {
                    'success': False,
                    'message': data.get('message', 'Failed to fetch mahasiswa')
                }
            
            mahasiswa_list = data.get('data', [])
            logger.info(f"Fetched {len(mahasiswa_list)} mahasiswa records")
            
            return {
                'success': True,
                'data': mahasiswa_list,
                'count': len(mahasiswa_list)
            }
            
        except requests.exceptions.Timeout:
            logger.error(f"Timeout fetching mahasiswa from {self.base_url}")
            return {
                'success': False,
                'message': 'Timeout: Server tidak merespon. Coba lagi atau periksa koneksi internet.'
            }
        except requests.exceptions.ConnectionError as e:
            logger.error(f"Connection error fetching mahasiswa: {e}")
            return {
                'success': False,
                'message': f'Gagal terhubung ke server: {str(e)}'
            }
        except requests.exceptions.HTTPError as e:
            logger.error(f"HTTP error fetching mahasiswa: {e}")
            return {
                'success': False,
                'message': f'Error HTTP {e.response.status_code}: Endpoint mungkin tidak tersedia'
            }
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching mahasiswa: {e}")
            return {
                'success': False,
                'message': f'Error fetching mahasiswa: {str(e)}'
            }
    
    def fetch_users(self) -> Dict:
        """
        Menarik semua data users (untuk akun login mahasiswa) dari Laravel API
        
        Returns:
            Dict dengan status dan data users
        """
        try:
            logger.info("Fetching users data from Laravel...")
            response = self.session.get(
                f"{self.base_url}/api/sync/users", 
                timeout=60,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            
            data = response.json()
            if not data.get('success'):
                return {
                    'success': False,
                    'message': data.get('message', 'Failed to fetch users')
                }
            
            users_list = data.get('data', [])
            logger.info(f"Fetched {len(users_list)} user records")
            
            return {
                'success': True,
                'data': users_list,
                'count': len(users_list)
            }
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching users: {e}")
            return {
                'success': False,
                'message': f'Error fetching users: {str(e)}'
            }
    
    def fetch_schedules(self) -> Dict:
        """
        Menarik data jadwal PKKMB dari Laravel API
        
        Returns:
            Dict dengan status dan data jadwal
        """
        try:
            logger.info("Fetching schedules from Laravel...")
            response = self.session.get(
                f"{self.base_url}/api/sync/schedules", 
                timeout=30,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            
            data = response.json()
            if not data.get('success'):
                return {
                    'success': False,
                    'message': data.get('message', 'Failed to fetch schedules')
                }
            
            schedules = data.get('data', [])
            logger.info(f"Fetched {len(schedules)} schedule records")
            
            return {
                'success': True,
                'data': schedules,
                'count': len(schedules)
            }
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching schedules: {e}")
            return {
                'success': False,
                'message': f'Error fetching schedules: {str(e)}'
            }
    
    def fetch_kegiatan(self) -> Dict:
        """
        Menarik data kegiatan dari Laravel API
        
        Returns:
            Dict dengan status dan data kegiatan
        """
        try:
            logger.info("Fetching kegiatan from Laravel...")
            response = self.session.get(
                f"{self.base_url}/api/sync/kegiatan", 
                timeout=30,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            
            data = response.json()
            if not data.get('success'):
                return {
                    'success': False,
                    'message': data.get('message', 'Failed to fetch kegiatan')
                }
            
            kegiatan_list = data.get('data', [])
            logger.info(f"Fetched {len(kegiatan_list)} kegiatan records")
            
            return {
                'success': True,
                'data': kegiatan_list,
                'count': len(kegiatan_list)
            }
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching kegiatan: {e}")
            return {
                'success': False,
                'message': f'Error fetching kegiatan: {str(e)}'
            }
    
    def fetch_system_config(self) -> Dict:
        """
        Menarik system config (termasuk grace period) dari Laravel API
        
        Returns:
            Dict dengan status dan data config
        """
        try:
            logger.info("Fetching system config from Laravel...")
            response = self.session.get(
                f"{self.base_url}/api/sync/system-config", 
                timeout=30,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            
            data = response.json()
            if not data.get('success'):
                return {
                    'success': False,
                    'message': data.get('message', 'Failed to fetch system config')
                }
            
            config_list = data.get('data', [])
            logger.info(f"Fetched {len(config_list)} system config records")
            
            return {
                'success': True,
                'data': config_list,
                'count': len(config_list)
            }
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching system config: {e}")
            return {
                'success': False,
                'message': f'Error fetching system config: {str(e)}'
            }
    
    def sync_mahasiswa_to_local(self, mahasiswa_list: List[Dict]) -> Dict:
        """
        Simpan/update data mahasiswa ke database MySQL lokal
        OTOMATIS MEMBUAT USER ACCOUNT untuk setiap mahasiswa
        
        Args:
            mahasiswa_list: List of mahasiswa data
            
        Returns:
            Dict dengan statistik sinkronisasi dan daftar akun yang gagal terbuat
        """
        inserted = 0
        updated = 0
        errors = 0
        users_created = 0
        users_updated = 0
        uncreated_users = []
        
        # Use single connection for batch sync to maximize performance & stability
        conn = self._get_conn()
        cursor = conn.cursor(dictionary=True)
        
        try:
            for mhs in mahasiswa_list:
                mhs_id = mhs.get('id')
                if not mhs_id:
                    continue
                    
                mhs_name = mhs.get('name', 'Tanpa Nama')
                is_active = 1 if mhs.get('is_active', True) else 0
                qr_code = mhs.get('qr_code_id') or mhs_id
                
                try:
                    # Check if mahasiswa exists for stats
                    cursor.execute("SELECT id FROM mahasiswa WHERE id = %s", (mhs_id,))
                    existing = cursor.fetchone()
                    
                    tgl_lahir = mhs.get('tanggal_lahir')
                    if tgl_lahir and 'T' in str(tgl_lahir):
                        tgl_lahir = str(tgl_lahir).split('T')[0]
                    elif not tgl_lahir:
                        tgl_lahir = None

                    cursor.execute("""
                        INSERT INTO mahasiswa (
                            id, name, kompi, jurusan, prodi, tanggal_lahir, email,
                            no_telp_mahasiswa, no_telp_ortu, qr_code_id, is_active
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                            name = VALUES(name),
                            kompi = VALUES(kompi),
                            jurusan = VALUES(jurusan),
                            prodi = VALUES(prodi),
                            tanggal_lahir = VALUES(tanggal_lahir),
                            email = VALUES(email),
                            no_telp_mahasiswa = VALUES(no_telp_mahasiswa),
                            no_telp_ortu = VALUES(no_telp_ortu),
                            qr_code_id = VALUES(qr_code_id),
                            is_active = VALUES(is_active)
                    """, (
                        mhs_id,
                        mhs_name,
                        mhs.get('kompi'),
                        mhs.get('jurusan'),
                        mhs.get('prodi'),
                        tgl_lahir,
                        mhs.get('email'),
                        mhs.get('no_telp_mahasiswa'),
                        mhs.get('no_telp_ortu'),
                        qr_code,
                        is_active
                    ))

                    if existing:
                        updated += 1
                    else:
                        inserted += 1
                    
                    # OTOMATIS BUAT/UPDATE USER ACCOUNT
                    user_res = self._sync_user_with_cursor(cursor, mhs)
                    if user_res.get('created'):
                        users_created += 1
                    elif user_res.get('updated'):
                        users_updated += 1
                    else:
                        uncreated_users.append({
                            'id': mhs_id,
                            'name': mhs_name,
                            'email': mhs.get('email', '-'),
                            'reason': user_res.get('error', 'Gagal membuat/update akun user')
                        })
                        
                except Exception as inner_e:
                    logger.error(f"Error syncing mahasiswa {mhs_id}: {inner_e}")
                    errors += 1
                    uncreated_users.append({
                        'id': mhs_id,
                        'name': mhs_name,
                        'email': mhs.get('email', '-'),
                        'reason': f"Database error: {str(inner_e)}"
                    })
                    
            conn.commit()
        except Exception as batch_e:
            logger.error(f"Batch sync error: {batch_e}")
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()
        
        return {
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'errors': errors,
            'total': len(mahasiswa_list),
            'users_created': users_created,
            'users_updated': users_updated,
            'uncreated_users': uncreated_users,
            'uncreated_count': len(uncreated_users)
        }
    
    def _sync_user_with_cursor(self, cursor, mahasiswa: Dict) -> Dict:
        """
        Buat atau update user account untuk mahasiswa menggunakan cursor yang sama
        """
        try:
            username = str(mahasiswa['id']).strip()
            full_name = mahasiswa.get('name', '')
            is_active = 1 if mahasiswa.get('is_active', True) else 0
            
            kompi = mahasiswa.get('kompi')
            
            if not username:
                return {'created': False, 'updated': False, 'error': 'ID / Username kosong'}
                
            cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
            existing_user = cursor.fetchone()
            
            import hashlib
            default_password = hashlib.md5(username.encode()).hexdigest()
            
            cursor.execute("""
                INSERT INTO users (username, password, full_name, role, is_active, assigned_kompi, mahasiswa_id)
                VALUES (%s, %s, %s, 'mahasiswa', %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    full_name = VALUES(full_name),
                    is_active = VALUES(is_active),
                    assigned_kompi = VALUES(assigned_kompi),
                    mahasiswa_id = VALUES(mahasiswa_id),
                    role = 'mahasiswa'
            """, (username, default_password, full_name, is_active, kompi, username))

            if existing_user:
                return {'created': False, 'updated': True}
            else:
                return {'created': True, 'updated': False}
                
        except Exception as e:
            logger.error(f"Error syncing user for {mahasiswa.get('id')}: {e}")
            return {'created': False, 'updated': False, 'error': str(e)}
            
    def _sync_user_for_mahasiswa(self, mahasiswa: Dict) -> Dict:
        """
        Buat atau update user account untuk mahasiswa (standalone method)
        """
        conn = self._get_conn()
        cursor = conn.cursor(dictionary=True)
        try:
            res = self._sync_user_with_cursor(cursor, mahasiswa)
            conn.commit()
            return res
        except Exception as e:
            conn.rollback()
            return {'created': False, 'updated': False, 'error': str(e)}
        finally:
            cursor.close()
            conn.close()
    
    def sync_schedules_to_local(self, schedules: List[Dict]) -> Dict:
        """
        Simpan/update data jadwal PKKMB ke database MySQL lokal
        
        Args:
            schedules: List of schedule data
            
        Returns:
            Dict dengan statistik sinkronisasi
        """
        inserted = 0
        updated = 0
        errors = 0
        
    def _parse_date_string(self, tanggal_str):
        """Parse string tanggal ke object date (datetime.date) dengan penanganan timezone"""
        if not tanggal_str:
            return None
        try:
            from dateutil import parser as dateparser
            import pytz
            
            # Coba isoparse dulu
            try:
                dt = dateparser.isoparse(tanggal_str)
            except Exception:
                dt = dateparser.parse(tanggal_str)
                
            if dt.tzinfo is not None:
                jakarta_tz = pytz.timezone('Asia/Jakarta')
                dt = dt.astimezone(jakarta_tz)
            return dt.date()
        except Exception as e:
            logger.warning(f"Gagal parse date '{tanggal_str}': {e}")
            if isinstance(tanggal_str, str) and len(tanggal_str) >= 10:
                return tanggal_str[:10]
            return tanggal_str

    def sync_schedules_to_local(self, schedules: List[Dict]) -> Dict:
        """
        Simpan/update data jadwal PKKMB ke database MySQL lokal
        
        Args:
            schedules: List of schedule data
            
        Returns:
            Dict dengan statistik sinkronisasi
        """
        inserted = 0
        updated = 0
        errors = 0
        
        for schedule in schedules:
            try:
                tanggal_str = schedule.get('tanggal')
                tanggal_date = self._parse_date_string(tanggal_str)
                is_active = 1 if schedule.get('is_active', True) else 0
                
                # Check if schedule exists
                existing = self._execute(
                    "SELECT id FROM pkkmb_schedules WHERE id = %s",
                    (schedule['id'],),
                    fetch_one=True
                )
                
                if existing:
                    # Update existing record
                    self._execute("""
                        UPDATE pkkmb_schedules SET
                            hari_ke = %s,
                            tanggal = %s,
                            check_in_start = %s,
                            check_in_end = %s,
                            check_out_start = %s,
                            check_out_end = %s,
                            is_active = %s
                        WHERE id = %s
                    """, (
                        schedule.get('hari_ke'),
                        tanggal_date,
                        schedule.get('check_in_start'),
                        schedule.get('check_in_end'),
                        schedule.get('check_out_start'),
                        schedule.get('check_out_end'),
                        is_active,
                        schedule['id']
                    ))
                    updated += 1
                else:
                    # Insert new record
                    self._execute("""
                        INSERT INTO pkkmb_schedules (
                            id, hari_ke, tanggal, check_in_start, check_in_end,
                            check_out_start, check_out_end, is_active
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        schedule['id'],
                        schedule.get('hari_ke'),
                        tanggal_date,
                        schedule.get('check_in_start'),
                        schedule.get('check_in_end'),
                        schedule.get('check_out_start'),
                        schedule.get('check_out_end'),
                        is_active
                    ))
                    inserted += 1
                    
            except Exception as e:
                logger.error(f"Error syncing schedule {schedule.get('id')}: {e}")
                errors += 1
        
        return {
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'errors': errors,
            'total': len(schedules)
        }
    
    def sync_kegiatan_to_local(self, kegiatan_list: List[Dict]) -> Dict:
        """
        Simpan/update data kegiatan ke database MySQL lokal
        
        Args:
            kegiatan_list: List of kegiatan data
            
        Returns:
            Dict dengan statistik sinkronisasi
        """
        inserted = 0
        updated = 0
        errors = 0
        
        for kegiatan in kegiatan_list:
            try:
                tanggal_str = kegiatan.get('tanggal_pelaksanaan')
                tanggal_date = self._parse_date_string(tanggal_str)
                is_active = 1 if kegiatan.get('is_active', True) else 0
                
                # Check if kegiatan exists
                existing = self._execute(
                    "SELECT id FROM kegiatan WHERE id = %s",
                    (kegiatan['id'],),
                    fetch_one=True
                )
                
                if existing:
                    # Update existing record
                    self._execute("""
                        UPDATE kegiatan SET
                            nama = %s,
                            tanggal_pelaksanaan = %s,
                            is_active = %s
                        WHERE id = %s
                    """, (
                        kegiatan.get('nama'),
                        tanggal_date,
                        is_active,
                        kegiatan['id']
                    ))
                    updated += 1
                else:
                    # Insert new record
                    self._execute("""
                        INSERT INTO kegiatan (id, nama, tanggal_pelaksanaan, is_active)
                        VALUES (%s, %s, %s, %s)
                    """, (
                        kegiatan['id'],
                        kegiatan.get('nama'),
                        tanggal_date,
                        is_active
                    ))
                    inserted += 1
                    
            except Exception as e:
                logger.error(f"Error syncing kegiatan {kegiatan.get('id')}: {e}")
                errors += 1
        
        return {
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'errors': errors,
            'total': len(kegiatan_list)
        }
    
    def sync_system_config_to_local(self, config_list: List[Dict]) -> Dict:
        """
        Simpan/update system config ke database MySQL lokal
        
        Args:
            config_list: List of config data
            
        Returns:
            Dict dengan statistik sinkronisasi
        """
        inserted = 0
        updated = 0
        errors = 0
        
        for config in config_list:
            try:
                config_key = config.get('config_key')
                if not config_key:
                    logger.warning("Skipping system config item with empty config_key")
                    continue
                    
                config_value = config.get('config_value')
                description = config.get('description', '')
                
                # Check if config exists
                existing = self._execute(
                    "SELECT id FROM system_config WHERE config_key = %s",
                    (config_key,),
                    fetch_one=True
                )
                
                if existing:
                    # Update existing record
                    self._execute("""
                        UPDATE system_config SET
                            config_value = %s,
                            description = %s
                        WHERE config_key = %s
                    """, (config_value, description, config_key))
                    updated += 1
                else:
                    # Insert new record
                    self._execute("""
                        INSERT INTO system_config (config_key, config_value, description)
                        VALUES (%s, %s, %s)
                    """, (config_key, config_value, description))
                    inserted += 1
                    
            except Exception as e:
                logger.error(f"Error syncing system config {config.get('config_key')}: {e}")
                errors += 1
        
        return {
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'errors': errors,
            'total': len(config_list)
        }
    
    def sync_all(self) -> Dict:
        """
        Sinkronisasi semua data dari Laravel (mahasiswa, schedules, kegiatan, system_config)
        Otomatis membuat user account untuk setiap mahasiswa
        
        Returns:
            Dict dengan status sinkronisasi lengkap
        """
        results = {
            'success': True,
            'mahasiswa': {'success': False},
            'schedules': {'success': False},
            'kegiatan': {'success': False},
            'system_config': {'success': False}
        }
        
        # Test connection first
        conn_test = self.test_connection()
        if not conn_test['success']:
            results['success'] = False
            results['message'] = conn_test['message']
            return results
        
        # Sync Mahasiswa (+ auto create users)
        logger.info("Starting mahasiswa sync...")
        mhs_fetch = self.fetch_mahasiswa()
        if mhs_fetch['success']:
            mhs_sync = self.sync_mahasiswa_to_local(mhs_fetch['data'])
            results['mahasiswa'] = mhs_sync
        else:
            results['mahasiswa'] = mhs_fetch
            results['success'] = False
        
        # Sync Schedules
        logger.info("Starting schedules sync...")
        sch_fetch = self.fetch_schedules()
        if sch_fetch['success']:
            sch_sync = self.sync_schedules_to_local(sch_fetch['data'])
            results['schedules'] = sch_sync
        else:
            results['schedules'] = sch_fetch
            results['success'] = False
        
        # Sync Kegiatan
        logger.info("Starting kegiatan sync...")
        keg_fetch = self.fetch_kegiatan()
        if keg_fetch['success']:
            keg_sync = self.sync_kegiatan_to_local(keg_fetch['data'])
            results['kegiatan'] = keg_sync
        else:
            results['kegiatan'] = keg_fetch
            results['success'] = False
        
        # Sync System Config
        logger.info("Starting system config sync...")
        cfg_fetch = self.fetch_system_config()
        if cfg_fetch['success']:
            cfg_sync = self.sync_system_config_to_local(cfg_fetch['data'])
            results['system_config'] = cfg_sync
        else:
            results['system_config'] = cfg_fetch
            # Don't fail completely if system_config sync fails
            logger.warning("System config sync failed, but continuing...")
        
        # Sync Attendance from Laravel
        logger.info("Starting attendance sync from Laravel...")
        att_fetch = self.fetch_attendance()
        if att_fetch['success']:
            att_sync = self.sync_attendance_to_local(att_fetch['data'])
            results['attendance'] = att_sync
        else:
            results['attendance'] = att_fetch
            logger.warning("Attendance sync from Laravel failed, but continuing...")

        logger.info("Sync completed!")
        return results

    def fetch_attendance(self) -> Dict:
        """
        Menarik data absensi hari ini dari Laravel API
        """
        try:
            logger.info("Fetching attendance from Laravel...")
            response = self.session.get(
                f"{self.base_url}/api/sync/attendance", 
                timeout=30,
                verify=self.verify_ssl
            )
            response.raise_for_status()
            
            data = response.json()
            if not data.get('success'):
                return {
                    'success': False,
                    'message': data.get('message', 'Failed to fetch attendance')
                }
            
            attendances = data.get('data', [])
            logger.info(f"Fetched {len(attendances)} attendance records")
            
            return {
                'success': True,
                'data': attendances,
                'count': len(attendances)
            }
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching attendance: {e}")
            return {
                'success': False,
                'message': f'Error fetching attendance: {str(e)}'
            }

    def _format_datetime_for_mysql(self, val, date_context=None):
        """Format string/ISO/datetime object to MySQL DATETIME format (YYYY-MM-DD HH:MM:SS).
        Otomatis konversi UTC → WIB (UTC+7) jika timestamp mengandung timezone indicator Z atau +00:00."""
        if not val:
            return None
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ('none', 'null', '', '-', 'manual', 'usb-scanner', 'web-scanner', 'n/a'):
            return None

        import re
        from datetime import datetime, timezone, timedelta

        # Support time-only strings like '06.34', '06:34', '06.34.12', '06:34:12'
        time_match = re.match(r'^(\d{1,2})[:.](\d{2})(?:[:.](\d{2}))?$', val_str)
        if time_match:
            hh, mm, ss = int(time_match.group(1)), int(time_match.group(2)), int(time_match.group(3) or 0)
            date_part = date_context or datetime.now().strftime('%Y-%m-%d')
            return f"{date_part} {hh:02d}:{mm:02d}:{ss:02d}"

        try:
            from dateutil import parser as dateparser
            dt = dateparser.parse(val_str, dayfirst=True)
            if dt.tzinfo is not None:
                wib = timezone(timedelta(hours=7))
                dt = dt.astimezone(wib)
            elif val_str.endswith('Z') or '+00:00' in val_str:
                from datetime import timezone as tz, timedelta as td
                dt = dt.replace(tzinfo=tz.utc).astimezone(tz(td(hours=7)))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception as e:
            logger.warning(f"Failed to parse datetime '{val}': {e}")
            return None

    def _format_date_for_mysql(self, date_val, check_in=None):
        """Format date string to MySQL DATE format (YYYY-MM-DD).
        Otomatis konversi UTC → WIB agar tanggal benar (misal 2026-08-09T17:00:00Z → 2026-08-10 di WIB)."""
        if date_val:
            try:
                val_str = str(date_val).strip()
                if val_str and val_str.lower() not in ('none', 'null', '', '-', 'manual', 'usb-scanner', 'web-scanner', 'n/a'):
                    from dateutil import parser as dateparser
                    from datetime import timezone, timedelta
                    dt = dateparser.parse(val_str, dayfirst=True)
                    if dt.tzinfo is not None:
                        wib = timezone(timedelta(hours=7))
                        dt = dt.astimezone(wib)
                    elif val_str.endswith('Z') or '+00:00' in val_str:
                        from datetime import timezone as tz, timedelta as td
                        dt = dt.replace(tzinfo=tz.utc).astimezone(tz(td(hours=7)))
                    return dt.strftime('%Y-%m-%d')
            except Exception:
                pass
        
        if check_in:
            formatted_ci = self._format_datetime_for_mysql(check_in)
            if formatted_ci and len(formatted_ci) >= 10:
                return formatted_ci[:10]
                
        from datetime import datetime
        return datetime.now().strftime('%Y-%m-%d')

    def _ensure_attendance_schema(self, cursor):
        """Pastikan semua kolom pendukung (is_synced, is_late, dll) selalu ada di tabel attendance (auto-heal jika terkena migrate:fresh)"""
        cols = [
            ("check_in_time", "TIME NULL"),
            ("check_out_time", "TIME NULL"),
            ("kegiatan_id", "BIGINT UNSIGNED NULL"),
            ("sesi_id", "BIGINT UNSIGNED NULL"),
            ("is_late", "TINYINT(1) DEFAULT 0"),
            ("late_duration", "INT DEFAULT 0"),
            ("is_synced", "TINYINT(1) DEFAULT 0")
        ]
        for col_name, col_type in cols:
            try:
                cursor.execute(f"ALTER TABLE attendance ADD COLUMN {col_name} {col_type}")
            except Exception:
                pass  # Kolom sudah ada

    def sync_attendance_to_local(self, attendances: List[Dict]) -> Dict:
        """
        Simpan/update data absensi dari Laravel ke database MySQL lokal (Bulk Optimized & Resilient)
        """
        if not attendances:
            return {'success': True, 'inserted': 0, 'updated': 0, 'errors': 0, 'total': 0}

        inserted = 0
        updated = 0
        errors = 0
        
        conn = self._get_conn()
        cursor = conn.cursor(dictionary=True)
        try:
            # 1. Auto-heal skema tabel attendance (antisipasi jika user jalankan php artisan migrate:fresh)
            self._ensure_attendance_schema(cursor)
            
            # 2. Nonaktifkan FK checks untuk kecepatan & pencegahan error relasi
            cursor.execute("SET FOREIGN_KEY_CHECKS=0")
            
            # 3. Ambil data existing ke memori Python 1x saja (mengatur N+1 query problem)
            cursor.execute("SELECT id, mahasiswa_id, kegiatan_id, date, check_in, check_out FROM attendance")
            existing_rows = cursor.fetchall() or []
            existing_map = {}
            for r in existing_rows:
                k_keg = (str(r['mahasiswa_id']), r['kegiatan_id'], str(r['date']) if r['date'] else None)
                k_day = (str(r['mahasiswa_id']), None, str(r['date']) if r['date'] else None)
                existing_map[k_keg] = r
                existing_map[k_day] = r

            for att in attendances:
                try:
                    raw_mhs_id = att.get('mahasiswa_id')
                    if not raw_mhs_id:
                        continue
                    mhs_id = str(raw_mhs_id).strip()
                    
                    check_in = self._format_datetime_for_mysql(att.get('check_in'))
                    check_out = self._format_datetime_for_mysql(att.get('check_out'))
                    date_val = self._format_date_for_mysql(att.get('date'), check_in)
                    kegiatan_id = att.get('kegiatan_id')
                    is_late = 1 if att.get('is_late') else 0
                    late_duration = att.get('late_duration', 0)
                    
                    status_raw = (att.get('status') or '').lower()
                    
                    # Jika data dari server belum ada check_in dan belum ada check_out,
                    # dan statusnya bukan sakit/izin, lewati (jangan simpan record kosong ke DB lokal)
                    if not check_in and not check_out and status_raw not in ['sakit', 'izin']:
                        continue

                    key_kegiatan = (mhs_id, kegiatan_id, date_val if not kegiatan_id else None)
                    key_daily = (mhs_id, None, date_val)
                    
                    existing = existing_map.get(key_kegiatan) or existing_map.get(key_daily)
                    status_val = att.get('status') or ('hadir' if check_in else 'alpha')
                    
                    if existing:
                        # Proteksi: Jika data lokal atau server berstatus 'sakit'/'izin', jangan ganti ke 'hadir' kecuali ada check_in & check_out lengkap
                        existing_status = (existing.get('status') or '').lower()
                        if (existing_status in ['sakit', 'izin'] or status_raw in ['sakit', 'izin']) and not (check_in and check_out):
                            status_val = status_raw if status_raw in ['sakit', 'izin'] else existing_status
                        
                        final_check_in = check_in if check_in else self._format_datetime_for_mysql(existing['check_in'])
                        final_check_out = check_out if check_out else self._format_datetime_for_mysql(existing['check_out'])
                        cursor.execute("""
                            UPDATE attendance SET
                                check_in = %s,
                                check_out = %s,
                                status = %s,
                                is_late = %s,
                                late_duration = %s,
                                is_synced = 1
                            WHERE id = %s
                        """, (final_check_in, final_check_out, status_val, is_late, late_duration, existing['id']))
                        updated += 1
                    else:
                        cursor.execute("""
                            INSERT INTO attendance (
                                mahasiswa_id, kegiatan_id, date, check_in, check_out,
                                status, is_late, late_duration, camera_id, is_synced
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'SYNC-LARAVEL', 1)
                        """, (mhs_id, kegiatan_id, date_val, check_in, check_out, status_val, is_late, late_duration))
                        inserted += 1
                except Exception as e:
                    logger.error(f"Error syncing attendance record for mhs {att.get('mahasiswa_id')}: {e}")
                    errors += 1
            
            # Commit sekaligus di akhir batch
            conn.commit()
            
        except Exception as e:
            logger.error(f"Error in batch attendance sync: {e}")
            conn.rollback()
            errors = len(attendances)
        finally:
            try:
                cursor.execute("SET FOREIGN_KEY_CHECKS=1")
            except Exception:
                pass
            cursor.close()
            conn.close()
            
        return {
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'errors': errors,
            'total': len(attendances)
        }

    def import_excel_backup_to_local(self, file_path_or_bytes) -> Dict:
        """
        Import data Mahasiswa & Absensi dari file Excel export Laravel (AttendanceExport / MahasiswaExport),
        termasuk file multi-sheet dan multi-tabel yang dikelompokkan per Prodi.
        Memasukkan data ke database MySQL lokal & memperbarui backup JSON internal.
        """
        import openpyxl
        import io
        
        mhs_inserted, mhs_updated = 0, 0
        att_inserted, att_updated = 0, 0
        errors = 0
        
        try:
            # Handle string path, FileStorage, or bytes
            if hasattr(file_path_or_bytes, 'read'):
                wb = openpyxl.load_workbook(filename=io.BytesIO(file_path_or_bytes.read()), data_only=True)
            elif isinstance(file_path_or_bytes, (bytes, bytearray)):
                wb = openpyxl.load_workbook(filename=io.BytesIO(file_path_or_bytes), data_only=True)
            else:
                wb = openpyxl.load_workbook(filename=file_path_or_bytes, data_only=True)
            
            conn = self._get_conn()
            cursor = conn.cursor(dictionary=True)
            self._ensure_attendance_schema(cursor)
            
            # Auto-widen column sizes if needed
            for alter_sql in [
                "ALTER TABLE mahasiswa MODIFY COLUMN id VARCHAR(100)",
                "ALTER TABLE attendance MODIFY COLUMN mahasiswa_id VARCHAR(100)",
                "ALTER TABLE users MODIFY COLUMN mahasiswa_id VARCHAR(100)"
            ]:
                try:
                    cursor.execute(alter_sql)
                except Exception:
                    pass

            try:
                cursor.execute("SET FOREIGN_KEY_CHECKS=0")
            except Exception:
                pass
            
            invalid_keywords = [
                'monitoring', 'absensi', 'live', 'periode', 'dicetak', 'total',
                'prodi:', 'jalur', 'nomor pendaftaran', 'nama mahasiswa',
                'status absensi', 'jam masuk', 'jam keluar', 'kamera'
            ]

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                col_map = {}
                
                for row in ws.iter_rows(values_only=True):
                    if not row or not any(row):
                        continue
                    
                    row_strs = [str(cell).strip() if cell is not None else '' for cell in row]
                    row_lowers = [s.lower() for s in row_strs]
                    
                    # Strict Header Detection: Row must contain 'nama' AND ('pendaftaran' or 'npm' or 'id')
                    has_nama = any('nama' in s for s in row_lowers)
                    has_id_col = any('pendaftaran' in s or 'npm' in s or s in ('id', 'no', 'mahasiswa_id') for s in row_lowers)

                    if has_nama and has_id_col:
                        new_col_map = {}
                        for idx, s in enumerate(row_lowers):
                            if 'pendaftaran' in s or 'npm' in s or s in ('id', 'mahasiswa_id'):
                                new_col_map['id'] = idx
                            elif 'nama' in s:
                                new_col_map['name'] = idx
                            elif 'email' in s:
                                new_col_map['email'] = idx
                            elif 'kompi' in s:
                                new_col_map['kompi'] = idx
                            elif 'jurusan' in s:
                                new_col_map['jurusan'] = idx
                            elif 'prodi' in s:
                                new_col_map['prodi'] = idx
                            elif 'tanggal' in s or 'date' in s:
                                new_col_map['date'] = idx
                            elif ('masuk' in s or 'check_in' in s or 'check in' in s) and not any(k in s for k in ['tipe', 'metode', 'status', 'kamera']):
                                new_col_map['check_in'] = idx
                            elif ('keluar' in s or 'check_out' in s or 'check out' in s) and not any(k in s for k in ['tipe', 'metode', 'status', 'kamera']):
                                new_col_map['check_out'] = idx
                            elif 'status' in s and 'scan' not in s:
                                new_col_map['status'] = idx

                        if 'name' in new_col_map or 'id' in new_col_map:
                            col_map = new_col_map
                            continue  # Skip header row itself

                    if not col_map:
                        continue  # No header established yet for this table section

                    mhs_id = row_strs[col_map['id']] if 'id' in col_map and col_map['id'] < len(row_strs) else ''
                    mhs_name = row_strs[col_map['name']] if 'name' in col_map and col_map['name'] < len(row_strs) else ''

                    # Skip empty / invalid / non-student rows
                    if not mhs_id and not mhs_name:
                        continue

                    # Filter out title banner & sub-header rows
                    mhs_id_lower = mhs_id.lower()
                    mhs_name_lower = mhs_name.lower()

                    if any(inv in mhs_id_lower for inv in invalid_keywords) or any(inv in mhs_name_lower for inv in invalid_keywords):
                        continue

                    if mhs_id_lower in ('no', 'nomor pendaftaran', 'npm', 'id', 'mahasiswa_id') or mhs_name_lower in ('nama mahasiswa', 'nama'):
                        continue

                    if not mhs_id and mhs_name:
                        mhs_id = mhs_name

                    # Truncate mhs_id to max 100 characters to prevent any MySQL length error
                    mhs_id = mhs_id[:100].strip()
                    mhs_name = mhs_name[:255].strip()

                    email = row_strs[col_map['email']] if 'email' in col_map and col_map['email'] < len(row_strs) else None
                    kompi = row_strs[col_map['kompi']] if 'kompi' in col_map and col_map['kompi'] < len(row_strs) else None
                    jurusan = row_strs[col_map['jurusan']] if 'jurusan' in col_map and col_map['jurusan'] < len(row_strs) else None
                    prodi = row_strs[col_map['prodi']] if 'prodi' in col_map and col_map['prodi'] < len(row_strs) else None

                    if email in ('-', 'nan', 'None', ''): email = None
                    if kompi in ('-', 'nan', 'None', ''): kompi = None
                    if jurusan in ('-', 'nan', 'None', ''): jurusan = None
                    if prodi in ('-', 'nan', 'None', ''): prodi = None

                    # 1. Update/Insert Mahasiswa
                    cursor.execute("SELECT id FROM mahasiswa WHERE id = %s", (mhs_id,))
                    existing_mhs = cursor.fetchone()

                    cursor.execute("""
                        INSERT INTO mahasiswa (
                            id, name, kompi, jurusan, prodi, email, qr_code_id, is_active
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, 1)
                        ON DUPLICATE KEY UPDATE
                            name = VALUES(name),
                            kompi = COALESCE(VALUES(kompi), kompi),
                            jurusan = COALESCE(VALUES(jurusan), jurusan),
                            prodi = COALESCE(VALUES(prodi), prodi),
                            email = COALESCE(VALUES(email), email)
                    """, (mhs_id, mhs_name or f"Mahasiswa ({mhs_id})", kompi, jurusan, prodi, email, mhs_id))

                    if existing_mhs:
                        mhs_updated += 1
                    else:
                        mhs_inserted += 1

                    # Auto sync user account
                    self._sync_user_with_cursor(cursor, {
                        'id': mhs_id,
                        'name': mhs_name or f"Mahasiswa ({mhs_id})",
                        'kompi': kompi,
                        'is_active': True
                    })

                    # 2. Extract Attendance data
                    date_raw = row_strs[col_map['date']] if 'date' in col_map and col_map['date'] < len(row_strs) else None
                    ci_raw = row_strs[col_map['check_in']] if 'check_in' in col_map and col_map['check_in'] < len(row_strs) else None
                    co_raw = row_strs[col_map['check_out']] if 'check_out' in col_map and col_map['check_out'] < len(row_strs) else None
                    status_raw = row_strs[col_map['status']] if 'status' in col_map and col_map['status'] < len(row_strs) else None

                    if date_raw in ('-', 'nan', 'None', ''): date_raw = None
                    if ci_raw in ('-', 'nan', 'None', ''): ci_raw = None
                    if co_raw in ('-', 'nan', 'None', ''): co_raw = None
                    if status_raw in ('-', 'nan', 'None', ''): status_raw = None

                    if date_raw or ci_raw or status_raw:
                        formatted_date = self._format_date_for_mysql(date_raw, ci_raw)
                        formatted_ci = self._format_datetime_for_mysql(ci_raw, date_context=formatted_date) if ci_raw else None
                        formatted_co = self._format_datetime_for_mysql(co_raw, date_context=formatted_date) if co_raw else None
                        status_clean = (status_raw or 'hadir').lower()

                        cursor.execute("SELECT id FROM attendance WHERE mahasiswa_id = %s AND date = %s", (mhs_id, formatted_date))
                        existing_att = cursor.fetchone()

                        if existing_att:
                            cursor.execute("""
                                UPDATE attendance SET
                                    check_in = COALESCE(%s, check_in),
                                    check_out = COALESCE(%s, check_out),
                                    status = %s,
                                    is_synced = 1
                                WHERE id = %s
                            """, (formatted_ci, formatted_co, status_clean, existing_att['id']))
                            att_updated += 1
                        else:
                            cursor.execute("""
                                INSERT INTO attendance (
                                    mahasiswa_id, date, check_in, check_out, status, camera_id, is_synced
                                ) VALUES (%s, %s, %s, %s, %s, 'IMPORT-EXCEL', 1)
                            """, (mhs_id, formatted_date, formatted_ci, formatted_co, status_clean))
                            att_inserted += 1

            conn.commit()
            try:
                cursor.execute("SET FOREIGN_KEY_CHECKS=1")
            except Exception:
                pass
            cursor.close()
            conn.close()

            return {
                'success': True,
                'message': f"Import Excel berhasil! Data Mahasiswa (+{mhs_inserted} / ~{mhs_updated}), Absensi (+{att_inserted} / ~{att_updated}).",
                'stats': {
                    'mhs_inserted': mhs_inserted,
                    'mhs_updated': mhs_updated,
                    'att_inserted': att_inserted,
                    'att_updated': att_updated
                }
            }
        except Exception as e:
            logger.error(f"Error importing Excel backup: {e}", exc_info=True)
            return {
                'success': False,
                'message': f"Gagal membaca file Excel: {str(e)}"
            }

