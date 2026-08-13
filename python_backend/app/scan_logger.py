"""
Scan Speed & Device Performance Logger for SIABSEN Python Backend.
Stores append-only scan logs in JSONL format (logs/scan_speed.jsonl)
and provides Excel export reports structured per day with 2 sheets (Scan Masuk & Scan Keluar).
"""

import os
import json
import time
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import logging

from app.device_info import get_device_info

logger = logging.getLogger(__name__)

# Base directory for logs
LOGS_DIR = Path(__file__).parent.parent / "logs"
LOGS_DIR.mkdir(parents=True, exist_ok=True)
JSONL_LOG_FILE = LOGS_DIR / "scan_speed.jsonl"

_log_lock = threading.Lock()

MONTH_NAMES_ID = {
    1: "januari", 2: "februari", 3: "maret", 4: "april",
    5: "mei", 6: "juni", 7: "juli", 8: "agustus",
    9: "september", 10: "oktober", 11: "november", 12: "desember"
}

def format_date_indonesian_filename(dt_or_str) -> str:
    """Format datetime or YYYY-MM-DD string to report_speed_DD_bulan_siabsensi_YYYY.xlsx"""
    if isinstance(dt_or_str, str):
        try:
            dt = datetime.strptime(dt_or_str[:10], "%Y-%m-%d")
        except Exception:
            dt = datetime.now()
    elif isinstance(dt_or_str, datetime):
        dt = dt_or_str
    else:
        dt = datetime.now()

    day = dt.day
    month_name = MONTH_NAMES_ID.get(dt.month, "bulan")
    year = dt.year
    return f"report_speed_{day}_{month_name}_siabsensi_{year}.xlsx"

def format_date_indonesian_label(dt_or_str) -> str:
    """Format datetime or YYYY-MM-DD string to '13 Agustus 2026'"""
    if isinstance(dt_or_str, str):
        try:
            dt = datetime.strptime(dt_or_str[:10], "%Y-%m-%d")
        except Exception:
            dt = datetime.now()
    elif isinstance(dt_or_str, datetime):
        dt = dt_or_str
    else:
        dt = datetime.now()

    day = dt.day
    month_name = MONTH_NAMES_ID.get(dt.month, "").capitalize()
    year = dt.year
    return f"{day} {month_name} {year}"

class ScanSpeedLogger:
    """Thread-safe append-only logger for scan speeds and device performance."""

    @staticmethod
    def log_scan(
        qr_code_id: str,
        mahasiswa_name: Optional[str] = None,
        status: str = "SUCCESS",
        scan_type: str = "masuk",  # "masuk" or "keluar"
        input_type: str = "USB-SCANNER",
        hardware_ms: float = 0.0,
        api_duration_ms: float = 0.0,
        client_ip: Optional[str] = None,
        custom_device_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Record a single scan event into the append-only JSONL log file."""
        device_info = get_device_info()
        now = datetime.now()

        hardware_ms = max(0.0, round(float(hardware_ms), 2))
        api_duration_ms = max(0.0, round(float(api_duration_ms), 2))
        total_duration_ms = round(hardware_ms + api_duration_ms, 2)
        total_duration_sec = round(total_duration_ms / 1000.0, 3)

        # Normalize scan_type
        scan_type_clean = "keluar" if "keluar" in str(scan_type).lower() or "out" in str(scan_type).lower() else "masuk"

        entry = {
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "date_str": now.strftime("%Y-%m-%d"),
            "iso_time": now.isoformat(),
            "brand": device_info.get("brand", "Unknown"),
            "model": device_info.get("model", "PC/Laptop"),
            "device_full": device_info.get("full_name", "Unknown Device"),
            "hostname": custom_device_name or device_info.get("hostname", "Local"),
            "os": device_info.get("os", "Unknown OS"),
            "client_ip": client_ip or "127.0.0.1",
            "input_type": input_type,
            "scan_type": scan_type_clean,
            "qr_code_id": str(qr_code_id),
            "mahasiswa_name": mahasiswa_name or f"Mahasiswa ({qr_code_id})",
            "status": status,
            "hardware_ms": hardware_ms,
            "api_duration_ms": api_duration_ms,
            "total_duration_ms": total_duration_ms,
            "total_duration_sec": total_duration_sec
        }

        # Append-only write (Thread-safe)
        try:
            with _log_lock:
                with open(JSONL_LOG_FILE, "a", encoding="utf-8") as f:
                    f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        except Exception as e:
            logger.error(f"Failed to append scan speed log: {e}")

        return entry

    @staticmethod
    def get_logs(limit: int = 10000, target_date: Optional[str] = None) -> List[Dict[str, Any]]:
        """Read scan logs, optionally filtered by YYYY-MM-DD date."""
        if not JSONL_LOG_FILE.exists():
            return []

        logs = []
        try:
            with _log_lock:
                with open(JSONL_LOG_FILE, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            try:
                                item = json.loads(line)
                                # Extract date string
                                item_date = item.get("date_str") or item.get("timestamp", "")[:10]
                                item["date_str"] = item_date
                                if target_date and item_date != target_date:
                                    continue
                                logs.append(item)
                            except Exception:
                                continue
        except Exception as e:
            logger.error(f"Failed to read scan logs: {e}")

        return logs[::-1][:limit]

    @staticmethod
    def get_available_report_dates() -> List[Dict[str, Any]]:
        """Get all distinct dates recorded in the logs with scan counts."""
        all_logs = ScanSpeedLogger.get_logs(limit=100000)
        date_map = {}

        for log in all_logs:
            d_str = log.get("date_str") or log.get("timestamp", "")[:10]
            if not d_str or len(d_str) < 10:
                continue

            if d_str not in date_map:
                date_map[d_str] = {
                    "date": d_str,
                    "filename": format_date_indonesian_filename(d_str),
                    "label": format_date_indonesian_label(d_str),
                    "total_scans": 0,
                    "masuk_count": 0,
                    "keluar_count": 0
                }

            date_map[d_str]["total_scans"] += 1
            st = log.get("scan_type", "masuk")
            if st == "keluar":
                date_map[d_str]["keluar_count"] += 1
            else:
                date_map[d_str]["masuk_count"] += 1

        # If empty, add today as default
        today_str = datetime.now().strftime("%Y-%m-%d")
        if not date_map:
            date_map[today_str] = {
                "date": today_str,
                "filename": format_date_indonesian_filename(today_str),
                "label": format_date_indonesian_label(today_str),
                "total_scans": 0,
                "masuk_count": 0,
                "keluar_count": 0
            }

        # Sort latest date first
        sorted_dates = sorted(date_map.values(), key=lambda x: x["date"], reverse=True)
        return sorted_dates

    @staticmethod
    def generate_excel_report(target_date: Optional[str] = None) -> Path:
        """
        Generate Excel report for 1 specific date.
        Filename format: report_speed_13_agustus_siabsensi_2026.xlsx
        Contains 2 Sheets:
          - Sheet 1: Scan Masuk
          - Sheet 2: Scan Keluar
        Matches Laravel AttendanceExport design theme!
        """
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not target_date:
            target_date = datetime.now().strftime("%Y-%m-%d")

        filename = format_date_indonesian_filename(target_date)
        date_label = format_date_indonesian_label(target_date)
        output_file = LOGS_DIR / filename

        logs_for_date = ScanSpeedLogger.get_logs(limit=10000, target_date=target_date)

        masuk_logs = [l for l in logs_for_date if l.get("scan_type", "masuk") != "keluar"]
        keluar_logs = [l for l in logs_for_date if l.get("scan_type") == "keluar"]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Styling definitions matching Laravel AttendanceExport
        fill_row1 = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark Navy
        fill_row2 = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Electric Blue
        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Arial", size=10, color="000000")

        fill_status_success = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        font_status_success = Font(name="Arial", size=10, bold=True, color="15803D")

        fill_status_error = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        font_status_error = Font(name="Arial", size=10, bold=True, color="B91C1C")

        fill_status_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        font_status_warn = Font(name="Arial", size=10, bold=True, color="B45309")

        border_thin = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )
        border_header = Border(
            left=Side(style="thin", color="1E293B"),
            right=Side(style="thin", color="1E293B"),
            top=Side(style="thin", color="1E293B"),
            bottom=Side(style="thin", color="1E293B")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        col_widths = {
            "A": 6,   # No
            "B": 22,  # Waktu Scan
            "C": 18,  # Merek Laptop
            "D": 18,  # Model Laptop
            "E": 18,  # Nama Perangkat
            "F": 18,  # Sistem Operasi
            "G": 16,  # Tipe Input
            "H": 18,  # NPM / QR Code
            "I": 28,  # Nama Mahasiswa
            "J": 22,  # Kecepatan Alat (ms)
            "K": 20,  # Kecepatan API (ms)
            "L": 20,  # Total Waktu (ms)
            "M": 20,  # Total Waktu (Detik)
            "N": 16   # Status
        }

        headers_list = [
            "No", "Waktu Scan", "Merek Laptop", "Model Laptop", "Nama Perangkat",
            "Sistem Operasi", "Tipe Input", "NPM / QR Code", "Nama Mahasiswa",
            "Kecepatan Alat (ms)", "Kecepatan API (ms)", "Total Waktu (ms)",
            "Total Waktu (Detik)", "Status"
        ]

        def build_sheet(sheet_title: str, subtitle_text: str, log_data: List[Dict[str, Any]]):
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            total_cols = len(headers_list)
            last_col_letter = get_column_letter(total_cols)

            # Row 1: Title Banner
            ws.merge_cells(f"A1:{last_col_letter}1")
            c1 = ws["A1"]
            c1.value = f"REPORT KECEPATAN SCAN - {sheet_title.upper()} (SIABSEN)"
            c1.font = font_title
            c1.fill = fill_row1
            c1.alignment = align_center
            ws.row_dimensions[1].height = 36

            # Row 2: Subtitle Banner
            ws.merge_cells(f"A2:{last_col_letter}2")
            c2 = ws["A2"]
            c2.value = f"Tanggal: {date_label} | {subtitle_text} | Data Log Append-Only (Anti-Hapus)"
            c2.font = font_subtitle
            c2.fill = fill_row2
            c2.alignment = align_center
            ws.row_dimensions[2].height = 26

            # Row 3: Empty
            ws.row_dimensions[3].height = 12

            # Row 4: Header
            ws.row_dimensions[4].height = 28
            for c_idx, h_name in enumerate(headers_list, start=1):
                cell = ws.cell(row=4, column=c_idx, value=h_name)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_header

            # Data Rows
            if log_data:
                for r_idx, log in enumerate(log_data):
                    curr_row = 5 + r_idx
                    ws.row_dimensions[curr_row].height = 22
                    is_even = (r_idx % 2 == 1)

                    row_vals = [
                        r_idx + 1,
                        log.get("timestamp"),
                        log.get("brand"),
                        log.get("model"),
                        log.get("hostname"),
                        log.get("os"),
                        log.get("input_type"),
                        log.get("qr_code_id"),
                        log.get("mahasiswa_name"),
                        log.get("hardware_ms"),
                        log.get("api_duration_ms"),
                        log.get("total_duration_ms"),
                        log.get("total_duration_sec"),
                        log.get("status")
                    ]

                    for c_idx, val in enumerate(row_vals, start=1):
                        cell = ws.cell(row=curr_row, column=c_idx, value=val)
                        cell.font = font_data
                        cell.border = border_thin
                        if is_even:
                            cell.fill = fill_zebra

                        col_name = headers_list[c_idx - 1]
                        if col_name == "No":
                            cell.alignment = align_center
                        elif "Kecepatan" in col_name or "Total Waktu" in col_name:
                            cell.alignment = align_right
                            if "ms" in col_name:
                                cell.number_format = '#,##0.0 "ms"'
                            elif "Detik" in col_name:
                                cell.number_format = '0.000 "s"'
                        elif col_name in ["Waktu Scan", "Tipe Input", "NPM / QR Code", "Sistem Operasi"]:
                            cell.alignment = align_center
                        elif col_name == "Status":
                            cell.alignment = align_center
                            val_str = str(val).upper()
                            if "SUCCESS" in val_str or "CHECKED" in val_str:
                                cell.fill = fill_status_success
                                cell.font = font_status_success
                            elif "REJECT" in val_str or "ERROR" in val_str:
                                cell.fill = fill_status_error
                                cell.font = font_status_error
                            else:
                                cell.fill = fill_status_warn
                                cell.font = font_status_warn
                        else:
                            cell.alignment = align_left

            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        # Sheet 1: Scan Masuk
        build_sheet("Scan Masuk", f"Total Scan Masuk: {len(masuk_logs)} Record", masuk_logs)

        # Sheet 2: Scan Keluar
        build_sheet("Scan Keluar", f"Total Scan Keluar: {len(keluar_logs)} Record", keluar_logs)

        wb.save(output_file)
        return output_file
